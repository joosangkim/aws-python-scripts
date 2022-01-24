import boto3
import xlsxwriter
from openpyxl import load_workbook, Workbook

client = boto3.client('dynamodb')
arn = "arn:aws:dynamodb:ap-northeast-2:{}:table/{}"

excel_doc= load_workbook('targets.xlsx')
sheet = excel_doc['Shee2']

non_protected = []
for r in range(2, sheet.max_row + 1): # skip header
    if not sheet.cell(r,2).value :
        non_protected.append(sheet.cell(r,1).value)
# print(len(non_protected))
# print(non_protected)
# for table in non_protected:
#     print(table)
#     response = client.tag_resource(
#         ResourceArn=arn.format(table),
#         Tags=[
#             {
#                 'Key': 'protected',
#                 'Value': 'false'
#             },
#         ]
#     )
# all_tables = []
# res = client.list_tables()
# all_tables += res['TableNames']

# if res['LastEvaluatedTableName'] :
#     res = client.list_tables(
#         ExclusiveStartTableName=res['LastEvaluatedTableName']
#     )
#     all_tables += res['TableNames']

print("non_protected : {}".format(len(non_protected)))
removed = []
for t in non_protected:
    res = client.delete_table(
        TableName=t
    )
    print(res['TableDescription']['TableName'])
    removed.append(res['TableDescription']['TableName'])


# print("total : {}".format(len(targets)))
# for i in targets:
#     resp = client.list_tags_of_resource(
#         ResourceArn=arn.format(i)
#     )
#     for tag_dict in resp['Tags']:
#         if tag_dict['Key'] == 'protected':
#             if tag_dict['Value'] == 'true':
#                 use_table.append(i)
#         else:
#             continue

# use_table = list( dict.fromkeys(use_table) )
# print("used : {}".format(len(use_table)))
# for i in use_table:
#     targets.remove(i)
# targets = list( dict.fromkeys(targets) )
# print("removed : {}".format(len(targets)))

# wb = Workbook()
# ws = wb.active
# ws.append(['Table', 'Protected'])
# for use in use_table:
#     ws.append([use, True])
# for use in targets:
#     ws.append([use, False])

# wb.save("dynamos.xlsx")

    # print(resp['Tags'])
    # break
# for i in use_tables:
#     dblist.remove(i)

# for i in dblist:
#     res = client.list_tags_of_resource(
#         ResourceArn=arn.format(i)
#     )
#     targets.append(i)
    # if len(res['Tags']) == 0:
    #     targets.append(i)
    # else:
    #     tags = res['Tags']
    #     print(i, tags)
    #     for t in tags:
    #         if 'use' in t:
    #             print(i, t)
        #     print(t)
        #     if not ( t['Key'] == 'use' and t['Value'] == 'true'
        #         if :
        #             targets.append(i)

# print(targets)
# workbook = xlsxwriter.Workbook('targets.xlsx')
# worksheet = workbook.add_worksheet()
# r = 0
# c = 0
# for t in targets:
#     worksheet.write(
#         r, c, t
#     )
#     r+=1
# workbook.close()
# print(res)
# print(len(dblist))
# print(len(use_tables))
